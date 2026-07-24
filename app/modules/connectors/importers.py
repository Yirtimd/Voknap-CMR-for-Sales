import csv
import io
from pathlib import Path

from app.core.config import settings


TARGET_FIELDS = {
    "name",
    "phone",
    "email",
    "company_name",
    "lead_title",
    "source",
}

ALIASES = {
    "name": {"name", "contact", "contact_name", "фио", "контакт", "имя"},
    "phone": {"phone", "telephone", "телефон", "мобильный"},
    "email": {"email", "e-mail", "почта"},
    "company_name": {"company", "company_name", "organization", "компания", "организация"},
    "lead_title": {"lead", "lead_title", "title", "лид", "название лида"},
    "source": {"source", "источник"},
}


class ImportValidationError(ValueError):
    pass


def parse_import_file(filename: str, content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    if len(content) > settings.integration_import_max_bytes:
        raise ImportValidationError("Import file exceeds configured size limit")
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        rows = _parse_csv(content)
    elif suffix in {".xlsx", ".xlsm"}:
        rows = _parse_xlsx(content)
    else:
        raise ImportValidationError("Only CSV and XLSX files are supported")
    if not rows:
        raise ImportValidationError("Import file has no data rows")
    if len(rows) > settings.integration_import_max_rows:
        raise ImportValidationError("Import file exceeds configured row limit")
    headers = list(rows[0])
    return headers, rows


def suggest_mapping(headers: list[str]) -> dict[str, str]:
    result: dict[str, str] = {}
    for header in headers:
        normalized = header.strip().lower()
        for target, aliases in ALIASES.items():
            if normalized in aliases:
                result[header] = target
                break
    return result


def apply_mapping(rows: list[dict[str, str]], mapping: dict[str, str]) -> tuple[list[dict], list[dict]]:
    invalid_targets = sorted(set(mapping.values()) - TARGET_FIELDS)
    if invalid_targets:
        raise ImportValidationError(f"Unsupported target fields: {', '.join(invalid_targets)}")
    if "name" not in mapping.values() and "email" not in mapping.values():
        raise ImportValidationError("Map at least name or email")
    mapped: list[dict] = []
    errors: list[dict] = []
    for index, source in enumerate(rows, start=2):
        item = {
            target: str(source.get(column) or "").strip()
            for column, target in mapping.items()
        }
        if not item.get("name") and not item.get("email"):
            errors.append({"row": index, "message": "Name and email are empty"})
            continue
        mapped.append(item)
    return mapped, errors


def _parse_csv(content: bytes) -> list[dict[str, str]]:
    text = _decode_csv(content)
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    if not reader.fieldnames:
        raise ImportValidationError("CSV header is missing")
    return [
        {str(key).strip(): str(value or "") for key, value in row.items() if key}
        for row in reader
        if any(str(value or "").strip() for value in row.values())
    ]


def _parse_xlsx(content: bytes) -> list[dict[str, str]]:
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as error:
        raise ImportValidationError(f"Invalid XLSX file: {error}") from error
    sheet = workbook.active
    values = sheet.iter_rows(values_only=True)
    try:
        header_row = next(values)
    except StopIteration as error:
        raise ImportValidationError("XLSX file is empty") from error
    headers = [str(value or "").strip() for value in header_row]
    if not any(headers):
        raise ImportValidationError("XLSX header is missing")
    rows = []
    for values_row in values:
        if not any(value is not None and str(value).strip() for value in values_row):
            continue
        rows.append(
            {
                header: str(values_row[index] or "")
                for index, header in enumerate(headers)
                if header and index < len(values_row)
            }
        )
    workbook.close()
    return rows


def _decode_csv(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ImportValidationError("CSV encoding must be UTF-8 or Windows-1251")
